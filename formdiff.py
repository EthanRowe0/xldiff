#!/usr/bin/env python3
"""
FormDiff — Git-like line-by-line formula diffing for Excel workbooks.

Extracts raw formula definitions from two .xlsx files, serializes them into
a sorted linear array of  SheetName!Coordinate: =FORMULA  entries, and runs
a unified diff to surface additions, deletions, and modifications.
"""

import sys
import argparse
import difflib
import html as html_mod
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from colorama import init, Fore, Style

init(autoreset=True)

ARRAY_FORMULA_RE = re.compile(r"^\{(=.*)\}$", re.DOTALL)
COORD_RE = re.compile(r"\$?([A-Za-z]+)\$?(\d+)$")
ENTRY_COORD_RE = re.compile(r"^(.+)!(\$?[A-Za-z]+\$?\d+): ")
# Matches compact collapsed lines: -[×78] Sheet!A1:A78: =formula
COMPACT_ENTRY_RE = re.compile(r"^([+\-])\[×(\d+)\] (.+)!([A-Za-z0-9:$]+): (.+)$")

# Characters of equal context shown each side of a change before collapsing to ellipsis
_CTX_SHOW = 30
# Formula characters that make natural ellipsis boundaries (argument separators / grouping)
_FORMULA_BOUNDARIES = frozenset(',;()')


# ---------------------------------------------------------------------------
# Formula extraction
# ---------------------------------------------------------------------------

def _strip_formula_spaces(formula: str) -> str:
    """
    Remove all whitespace outside of double-quoted string literals.
    Spaces inside strings like "text value" are preserved.
    This ensures purely cosmetic spacing differences (indentation, spaces around
    commas/parens) are never treated as meaningful formula changes.
    """
    result: list[str] = []
    in_str = False
    for ch in formula:
        if ch == '"':
            in_str = not in_str
            result.append(ch)
        elif in_str or ch not in (' ', '\t'):
            result.append(ch)
    return ''.join(result)


def _normalize_formula(raw: str) -> str:
    """Strip array-formula braces, collapse embedded newlines, strip non-string spaces."""
    stripped = raw.strip()
    m = ARRAY_FORMULA_RE.match(stripped)
    if m:
        stripped = m.group(1)
    # Collapse embedded newlines (from Excel line-break formatting) to spaces first,
    # then strip all whitespace outside string literals so cosmetic spacing is ignored.
    collapsed = re.sub(r"[\r\n]+", " ", stripped).strip()
    return _strip_formula_spaces(collapsed)


def _coord_sort_key(entry: str) -> tuple:
    bang = entry.find("!")
    colon = entry.find(":", bang)
    if bang == -1 or colon == -1:
        return (entry,)
    sheet = entry[:bang]
    coord = entry[bang + 1 : colon]
    m = COORD_RE.search(coord)
    if not m:
        return (sheet, coord)
    col_letters = m.group(1).upper()
    row_num = int(m.group(2))
    col_index = 0
    for ch in col_letters:
        col_index = col_index * 26 + (ord(ch) - ord("A") + 1)
    return (sheet, col_index, row_num)


def extract_formulas(
    file_path: str,
    ignore_sheets: set[str] | None = None,
    only_sheets: set[str] | None = None,
) -> list[str]:
    """
    Open *file_path* in read-only mode and return a sorted list of formula
    strings in the format ``SheetName!A1: =SUM(B1:B10)``.
    """
    try:
        wb = load_workbook(file_path, data_only=False, read_only=True)
    except Exception as e:
        print(f"Error opening workbook {file_path}: {e}", file=sys.stderr)
        sys.exit(1)

    skipped = ignore_sheets or set()
    records: list[str] = []
    for sheet_name in wb.sheetnames:
        if sheet_name in skipped:
            continue
        if only_sheets is not None and sheet_name not in only_sheets:
            continue
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(values_only=False):
            for cell in row:
                val = cell.value
                if isinstance(val, str) and val.lstrip().startswith("="):
                    formula = _normalize_formula(val)
                    records.append(f"{sheet_name}!{cell.coordinate}: {formula}")

    wb.close()
    records.sort(key=_coord_sort_key)
    return records


# ---------------------------------------------------------------------------
# Entry parsing helpers
# ---------------------------------------------------------------------------

def _parse_entry(line: str) -> tuple[str, str, str, str] | None:
    """Parse a raw diff line into (sign, sheet, coord, formula)."""
    if not line or line.startswith("@@") or line.startswith("---") or line.startswith("+++"):
        return None
    sign = line[0]
    rest = line[1:]
    m = ENTRY_COORD_RE.match(rest)
    if not m:
        return None
    sheet = m.group(1)
    coord = m.group(2)
    formula = rest[m.end():]
    return (sign, sheet, coord, formula)


def _col_letters(coord: str) -> str:
    m = COORD_RE.search(coord.replace("$", ""))
    return m.group(1).upper() if m else ""


def _row_num(coord: str) -> int:
    m = COORD_RE.search(coord)
    return int(m.group(2)) if m else -1


def _col_to_index(col: str) -> int:
    idx = 0
    for ch in col.upper():
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return idx


# ---------------------------------------------------------------------------
# Compact diff
# ---------------------------------------------------------------------------

def build_compact_diff(diff_lines: list[str]) -> list[str]:
    """Collapse consecutive same-change runs into single annotated lines."""
    result: list[str] = []
    pending_removed: list[tuple] = []
    pending_added: list[tuple] = []

    def _flush(removed, added):
        out = []
        n_rem, n_add = len(removed), len(added)

        def is_collapsible(entries):
            if len(entries) < 2:
                return False
            sheet0 = entries[0][1]
            col0 = _col_letters(entries[0][2])
            formula0 = entries[0][3]
            for i, (_, sh, coord, fml, _) in enumerate(entries):
                if sh != sheet0 or _col_letters(coord) != col0 or fml != formula0:
                    return False
                if i > 0 and _row_num(entries[i][2]) != _row_num(entries[i-1][2]) + 1:
                    return False
            return True

        if n_rem == n_add and n_rem >= 2 and is_collapsible(removed) and is_collapsible(added):
            n = n_rem
            sheet = removed[0][1]
            col = _col_letters(removed[0][2])
            first_row = _row_num(removed[0][2])
            last_row  = _row_num(removed[-1][2])
            range_label = f"{sheet}!{col}{first_row}:{col}{last_row}"
            out.append(f"-[×{n}] {range_label}: {removed[0][3]}")
            out.append(f"+[×{n}] {range_label}: {added[0][3]}")
        else:
            for _, _, _, _, orig in removed:
                out.append(orig)
            for _, _, _, _, orig in added:
                out.append(orig)
        return out

    for line in diff_lines:
        parsed = _parse_entry(line)
        if parsed is None:
            result.extend(_flush(pending_removed, pending_added))
            pending_removed.clear()
            pending_added.clear()
            result.append(line)
            continue

        sign, sheet, coord, formula = parsed
        if sign == " ":
            result.extend(_flush(pending_removed, pending_added))
            pending_removed.clear()
            pending_added.clear()
            result.append(line)
        elif sign == "-":
            if pending_removed:
                prev = pending_removed[-1]
                if not (_col_letters(prev[2]) == _col_letters(coord) and prev[1] == sheet
                        and _row_num(coord) == _row_num(prev[2]) + 1 and prev[3] == formula):
                    result.extend(_flush(pending_removed, pending_added))
                    pending_removed.clear()
                    pending_added.clear()
            pending_removed.append((sign, sheet, coord, formula, line))
        elif sign == "+":
            if pending_added:
                prev = pending_added[-1]
                if not (_col_letters(prev[2]) == _col_letters(coord) and prev[1] == sheet
                        and _row_num(coord) == _row_num(prev[2]) + 1 and prev[3] == formula):
                    result.extend(_flush(pending_removed, pending_added))
                    pending_removed.clear()
                    pending_added.clear()
            pending_added.append((sign, sheet, coord, formula, line))

    result.extend(_flush(pending_removed, pending_added))
    return result


def _compact_output_path(output_path: str) -> str:
    p = Path(output_path)
    return str(p.with_stem(p.stem + "-compact"))


# ---------------------------------------------------------------------------
# Summary
# ---------------------------------------------------------------------------

def _parse_diff_entry(line: str) -> tuple[str, str, str, str] | None:
    parsed = _parse_entry(line)
    if parsed is None:
        return None
    return parsed


def _coords_to_ranges(coords: list[str]) -> str:
    if not coords:
        return ""
    parsed = []
    for c in coords:
        m = COORD_RE.search(c.replace("$", ""))
        if m:
            col = m.group(1).upper()
            row = int(m.group(2))
            parsed.append((_col_to_index(col), col, row))
    parsed.sort()
    if not parsed:
        return ", ".join(coords)
    ranges: list[str] = []
    run_col_idx, run_col, run_start, run_end = parsed[0]
    for col_idx, col, row in parsed[1:]:
        if col_idx == run_col_idx and row == run_end + 1:
            run_end = row
        else:
            ranges.append(f"{run_col}{run_start}" if run_end == run_start else f"{run_col}{run_start}:{run_col}{run_end}")
            run_col_idx, run_col, run_start, run_end = col_idx, col, row, row
    ranges.append(f"{run_col}{run_start}" if run_end == run_start else f"{run_col}{run_start}:{run_col}{run_end}")
    return ", ".join(ranges)


def print_summary(diff_lines: list[str]) -> None:
    removed: dict[str, list[tuple[str, str]]] = defaultdict(list)
    added:   dict[str, list[tuple[str, str]]] = defaultdict(list)
    for line in diff_lines:
        parsed = _parse_diff_entry(line)
        if parsed is None:
            continue
        sign, sheet, coord, formula = parsed
        if sign == "-":
            removed[sheet].append((coord, formula))
        elif sign == "+":
            added[sheet].append((coord, formula))

    all_sheets = sorted(set(list(removed) + list(added)))
    if not all_sheets:
        return

    print(Fore.CYAN + Style.BRIGHT + "\n─── Change Summary ───────────────────────────────────")
    for sheet in all_sheets:
        r, a = removed[sheet], added[sheet]
        n_removed, n_added = len(r), len(a)
        print(Style.BRIGHT + f"\n  {sheet}")
        if n_removed == 0:
            print(Fore.GREEN + f"    {n_added} formula(s) added  [{_coords_to_ranges([c for c,_ in a])}]")
        elif n_added == 0:
            print(Fore.RED + f"    {n_removed} formula(s) removed  [{_coords_to_ranges([c for c,_ in r])}]")
        else:
            pair_count = min(n_removed, n_added)
            pattern_coords: dict[tuple[str, str], list[str]] = defaultdict(list)
            for (coord, old_f), (_, new_f) in zip(r[:pair_count], a[:pair_count]):
                if old_f == new_f:
                    continue
                lo, hi_old, hi_new = 0, len(old_f), len(new_f)
                while lo < min(hi_old, hi_new) and old_f[lo] == new_f[lo]:
                    lo += 1
                while hi_old > lo and hi_new > lo and old_f[hi_old-1] == new_f[hi_new-1]:
                    hi_old -= 1; hi_new -= 1
                frm, to = old_f[lo:hi_old], new_f[lo:hi_new]
                if frm and to:
                    pattern_coords[(frm, to)].append(coord)
            print(f"    {max(n_removed, n_added)} formula(s) modified")
            for (frm, to), coords in sorted(pattern_coords.items(), key=lambda x: -len(x[1])):
                print(Fore.RED + f"      {len(coords)}x  {frm}" + Style.RESET_ALL + "  →  " + Fore.GREEN + f"{to}" + Style.DIM + f"   ({_coords_to_ranges(coords)})")
            net = n_added - n_removed
            if net != 0:
                print(f"    {abs(net)} formula(s) net {'added' if net > 0 else 'removed'}")
    print(Fore.CYAN + "──────────────────────────────────────────────────────\n")


# ---------------------------------------------------------------------------
# HTML report
# ---------------------------------------------------------------------------

def _boundary_cut_right(text: str, target: int) -> int:
    """
    Starting near *target*, scan forward up to 20 chars for a formula boundary
    character (comma, paren) and return the index just after it — so the boundary
    character itself is included in the visible head.
    Falls back to *target* if none found.
    """
    for i in range(target, min(target + 20, len(text))):
        if text[i] in _FORMULA_BOUNDARIES:
            return i + 1
    return target


def _boundary_cut_left(text: str, target: int) -> int:
    """
    Starting near *target*, scan backward up to 20 chars for a formula boundary
    character and return the index just after it — so the tail starts cleanly
    after a separator.
    Falls back to *target* if none found.
    """
    for i in range(min(target, len(text) - 1), max(target - 20, -1), -1):
        if text[i] in _FORMULA_BOUNDARIES:
            return i + 1
    return target


def _collapse_ctx(text: str) -> str:
    """
    Render an equal (unchanged) segment. If it exceeds _CTX_SHOW*2 characters,
    show the head and tail with a clickable ellipsis in between.
    Cut points snap to the nearest comma/paren so the ellipsis lands between
    formula arguments rather than mid-word.
    """
    if len(text) <= _CTX_SHOW * 2 + 10:
        return html_mod.escape(text)

    head_end   = _boundary_cut_right(text, _CTX_SHOW)
    tail_start = _boundary_cut_left(text, len(text) - _CTX_SHOW)

    # Safety: if snapping caused overlap, just fall back to no collapsing
    if tail_start <= head_end:
        return html_mod.escape(text)

    head = html_mod.escape(text[:head_end])
    tail = html_mod.escape(text[tail_start:])
    full = html_mod.escape(text)
    return (
        f'<span class="ctx-wrap">'
        f'<span class="ctx-short">{head}'
        f'<button class="ctx-btn" title="Click to expand hidden text">…</button>'
        f'{tail}</span>'
        f'<span class="ctx-full">{full}</span>'
        f'</span>'
    )


def _char_diff_html(old: str, new: str) -> tuple[str, str]:
    """
    Return (old_html, new_html) with:
    - changed characters wrapped in <mark> tags
    - long unchanged segments collapsed to clickable ellipses
    """
    sm = difflib.SequenceMatcher(None, old, new, autojunk=False)
    old_parts, new_parts = [], []
    for op, i1, i2, j1, j2 in sm.get_opcodes():
        if op == "equal":
            old_parts.append(_collapse_ctx(old[i1:i2]))
            new_parts.append(_collapse_ctx(new[j1:j2]))
        elif op == "replace":
            old_parts.append(f'<mark class="del-char">{html_mod.escape(old[i1:i2])}</mark>')
            new_parts.append(f'<mark class="add-char">{html_mod.escape(new[j1:j2])}</mark>')
        elif op == "delete":
            old_parts.append(f'<mark class="del-char">{html_mod.escape(old[i1:i2])}</mark>')
        elif op == "insert":
            new_parts.append(f'<mark class="add-char">{html_mod.escape(new[j1:j2])}</mark>')
    return "".join(old_parts), "".join(new_parts)


def _parse_any_entry(line: str) -> tuple[str, str, str, str, int] | None:
    """
    Parse both regular and compact diff lines.
    Returns (sign, sheet, coord_label, formula, count) or None.
    count=1 for regular lines, count=N for compact [×N] lines.
    """
    # Try compact format first: -[×78] Sheet!A1:A78: =formula
    m = COMPACT_ENTRY_RE.match(line)
    if m:
        return (m.group(1), m.group(3), m.group(4), m.group(5), int(m.group(2)))
    # Try regular format
    parsed = _parse_entry(line)
    if parsed:
        sign, sheet, coord, formula = parsed
        return (sign, sheet, coord, formula, 1)
    return None


def _build_html_report(
    diff_lines: list[str],
    base_file: str,
    target_file: str,
) -> str:
    """Generate a self-contained HTML diff report from unified diff lines."""

    rows_by_sheet: dict[str, list[dict]] = defaultdict(list)
    # (coord, formula, count)
    pending_removed: list[tuple[str, str, int]] = []
    pending_added:   list[tuple[str, str, int]] = []

    def flush_pairs(sheet: str) -> None:
        pairs = min(len(pending_removed), len(pending_added))
        for i in range(pairs):
            coord_r, old_f, cnt_r = pending_removed[i]
            coord_a, new_f, cnt_a = pending_added[i]
            coord = coord_r if coord_r == coord_a else f"{coord_r} / {coord_a}"
            count = max(cnt_r, cnt_a)
            old_html, new_html = _char_diff_html(old_f, new_f)
            rows_by_sheet[sheet].append({
                "type": "change",
                "coord": html_mod.escape(coord),
                "count": count,
                "old": old_html,
                "new": new_html,
            })
        for coord, formula, count in pending_removed[pairs:]:
            rows_by_sheet[sheet].append({
                "type": "remove",
                "coord": html_mod.escape(coord),
                "count": count,
                "old": html_mod.escape(formula),
                "new": "",
            })
        for coord, formula, count in pending_added[pairs:]:
            rows_by_sheet[sheet].append({
                "type": "add",
                "coord": html_mod.escape(coord),
                "count": count,
                "old": "",
                "new": html_mod.escape(formula),
            })

    current_sheet = None
    for line in diff_lines:
        parsed = _parse_any_entry(line)
        if parsed is None:
            continue
        sign, sheet, coord, formula, count = parsed

        if sheet != current_sheet:
            if current_sheet is not None:
                flush_pairs(current_sheet)
                pending_removed.clear()
                pending_added.clear()
            current_sheet = sheet

        if sign == "-":
            if pending_added:
                flush_pairs(current_sheet)
                pending_removed.clear()
                pending_added.clear()
            pending_removed.append((coord, formula, count))
        elif sign == "+":
            pending_added.append((coord, formula, count))

    if current_sheet is not None:
        flush_pairs(current_sheet)

    # Build HTML
    CSS = """
    * { box-sizing: border-box; margin: 0; padding: 0; }
    body { font-family: ui-monospace, 'Cascadia Code', 'Fira Code', monospace;
           font-size: 12px; background: #0d1117; color: #c9d1d9; }
    h1 { font-size: 14px; font-weight: 600; padding: 16px 20px;
         background: #161b22; border-bottom: 1px solid #30363d; color: #e6edf3; }
    h1 span { color: #8b949e; font-weight: 400; }
    .sheet-block { margin: 16px 20px; border: 1px solid #30363d; border-radius: 6px; overflow: hidden; }
    .sheet-header { padding: 8px 14px; background: #161b22; border-bottom: 1px solid #30363d;
                    font-weight: 600; color: #e6edf3; cursor: pointer; user-select: none;
                    display: flex; justify-content: space-between; align-items: center; }
    .sheet-header:hover { background: #1c2128; }
    .sheet-count { font-size: 11px; font-weight: 400; color: #8b949e; }
    .sheet-body { display: block; }
    .sheet-body.collapsed { display: none; }
    table { width: 100%; border-collapse: collapse; }
    col.col-coord { width: 90px; }
    col.col-old, col.col-new { width: calc(50% - 45px); }
    thead th { padding: 6px 10px; background: #161b22; color: #8b949e;
               font-weight: 500; font-size: 11px; text-align: left;
               border-bottom: 1px solid #30363d; position: sticky; top: 0; }
    tr { border-bottom: 1px solid #21262d; }
    tr:last-child { border-bottom: none; }
    td { padding: 5px 10px; vertical-align: top; word-break: break-all; line-height: 1.6; }
    td.coord { color: #8b949e; white-space: nowrap; }
    .count-badge { display: inline-block; font-size: 10px; background: #30363d;
                   color: #8b949e; border-radius: 10px; padding: 0 5px; margin-top: 2px; }
    tr.change td.old { background: #3d1f1f; color: #ffa198; }
    tr.change td.new { background: #1a2d1a; color: #7ee787; }
    tr.remove td.old { background: #3d1f1f; color: #ffa198; }
    tr.remove td.new { background: transparent; }
    tr.add    td.old { background: transparent; }
    tr.add    td.new { background: #1a2d1a; color: #7ee787; }
    mark.del-char { background: #b62324; color: #fff; border-radius: 2px; padding: 0 1px; }
    mark.add-char { background: #1f6823; color: #fff; border-radius: 2px; padding: 0 1px; }
    .summary-bar { padding: 10px 20px 6px; color: #8b949e; font-size: 11px; }
    /* Ellipsis context collapsing */
    .ctx-full  { display: none; }
    .ctx-short { display: inline; }
    .ctx-btn   { background: none; border: 1px solid #444c56; border-radius: 3px;
                 color: #8b949e; cursor: pointer; font-size: 11px; padding: 0 4px;
                 margin: 0 2px; line-height: 1.4; vertical-align: middle; }
    .ctx-btn:hover { background: #30363d; color: #e6edf3; }
    /* Per-row expand/collapse button */
    .row-ctx-btn { display: block; margin-top: 5px; width: 100%; background: none;
                   border: 1px solid #30363d; border-radius: 3px; color: #8b949e;
                   cursor: pointer; font-size: 10px; padding: 1px 4px; text-align: center; }
    .row-ctx-btn:hover { background: #21262d; color: #e6edf3; }
    /* Global toolbar */
    .toolbar { padding: 8px 20px; background: #161b22; border-bottom: 1px solid #30363d;
               display: flex; gap: 8px; align-items: center; }
    .toolbar-btn { background: #21262d; border: 1px solid #30363d; border-radius: 4px;
                   color: #c9d1d9; cursor: pointer; font-size: 11px; padding: 3px 10px; }
    .toolbar-btn:hover { background: #30363d; color: #e6edf3; }
    """

    JS = """
    // ── Sheet collapse ──────────────────────────────────────────────────────
    document.querySelectorAll('.sheet-header').forEach(h => {
        h.addEventListener('click', e => {
            if (e.target.closest('.row-ctx-btn, .ctx-btn, .toolbar-btn')) return;
            const body = h.nextElementSibling;
            body.classList.toggle('collapsed');
            h.querySelector('.toggle').textContent =
                body.classList.contains('collapsed') ? '▶' : '▼';
        });
    });

    // ── Individual ellipsis expand ───────────────────────────────────────────
    function expandWrap(wrap) {
        wrap.querySelector('.ctx-short').style.display = 'none';
        wrap.querySelector('.ctx-full').style.display  = 'inline';
        wrap.dataset.expanded = '1';
    }
    function collapseWrap(wrap) {
        wrap.querySelector('.ctx-short').style.display = 'inline';
        wrap.querySelector('.ctx-full').style.display  = 'none';
        delete wrap.dataset.expanded;
    }
    document.querySelectorAll('.ctx-btn').forEach(btn => {
        btn.addEventListener('click', e => {
            e.stopPropagation();
            expandWrap(btn.closest('.ctx-wrap'));
        });
    });

    // ── Per-row expand / collapse ────────────────────────────────────────────
    document.querySelectorAll('.row-ctx-btn').forEach(btn => {
        btn.addEventListener('click', e => {
            e.stopPropagation();
            const row = btn.closest('tr');
            const wraps = row.querySelectorAll('.ctx-wrap');
            const anyCollapsed = [...wraps].some(w => !w.dataset.expanded);
            wraps.forEach(w => anyCollapsed ? expandWrap(w) : collapseWrap(w));
            btn.textContent = anyCollapsed ? '− collapse context' : '+ expand context';
        });
    });

    // ── Global expand / collapse all ────────────────────────────────────────
    let allExpanded = false;
    document.getElementById('btn-expand-all').addEventListener('click', () => {
        allExpanded = !allExpanded;
        document.querySelectorAll('.ctx-wrap').forEach(
            w => allExpanded ? expandWrap(w) : collapseWrap(w)
        );
        document.querySelectorAll('.row-ctx-btn').forEach(btn => {
            btn.textContent = allExpanded ? '− collapse context' : '+ expand context';
        });
        document.getElementById('btn-expand-all').textContent =
            allExpanded ? '− Collapse all context' : '+ Expand all context';
    });
    """

    sheet_blocks = []
    total_changes = sum(len(v) for v in rows_by_sheet.values())

    for sheet in sorted(rows_by_sheet):
        rows = rows_by_sheet[sheet]
        n = len(rows)
        table_rows = []
        for r in rows:
            count_badge = (
                f'<br><span class="count-badge">×{r["count"]}</span>'
                if r["count"] > 1 else ""
            )
            # Only show per-row button if the row actually has collapsible content
            has_ctx = 'ctx-wrap' in r.get("old", "") or 'ctx-wrap' in r.get("new", "")
            row_btn = '<button class="row-ctx-btn">+ expand context</button>' if has_ctx else ""
            coord_cell = f'<td class="coord">{r["coord"]}{count_badge}{row_btn}</td>'
            old_cell = (
                f'<td class="old">{r["old"]}</td>'
                if r["old"] != ""
                else '<td class="old" style="background:transparent"></td>'
            )
            new_cell = (
                f'<td class="new">{r["new"]}</td>'
                if r["new"] != ""
                else '<td class="new" style="background:transparent"></td>'
            )
            table_rows.append(
                f'<tr class="{r["type"]}">'
                f'{coord_cell}'
                f'{old_cell}{new_cell}'
                f'</tr>'
            )
        block = f"""
        <div class="sheet-block">
          <div class="sheet-header">
            <span>&#128196; {html_mod.escape(sheet)}</span>
            <span><span class="sheet-count">{n} change{"s" if n != 1 else ""}</span>&nbsp;&nbsp;<span class="toggle">▼</span></span>
          </div>
          <div class="sheet-body">
            <table>
              <colgroup><col class="col-coord"><col class="col-old"><col class="col-new"></colgroup>
              <thead><tr><th>Cell</th><th>Before</th><th>After</th></tr></thead>
              <tbody>{"".join(table_rows)}</tbody>
            </table>
          </div>
        </div>"""
        sheet_blocks.append(block)

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>FormDiff Report</title>
<style>{CSS}</style>
</head>
<body>
<h1>FormDiff Report &nbsp;<span>{html_mod.escape(base_file)} → {html_mod.escape(target_file)}</span></h1>
<div class="toolbar">
  <span style="color:#8b949e;font-size:11px">{total_changes} change{"s" if total_changes != 1 else ""} across {len(rows_by_sheet)} sheet{"s" if len(rows_by_sheet) != 1 else ""}</span>
  <button class="toolbar-btn" id="btn-expand-all">+ Expand all context</button>
</div>
{"".join(sheet_blocks)}
<script>{JS}</script>
</body>
</html>"""


def _html_output_path(output_path: str) -> str:
    return str(Path(output_path).with_suffix(".html"))


# ---------------------------------------------------------------------------
# Core diff runner
# ---------------------------------------------------------------------------

def run_diff(
    base_file: str,
    target_file: str,
    *,
    output_path: str | None = None,
    ignore_sheets: set[str] | None = None,
    only_sheets: set[str] | None = None,
    summary: bool = False,
    compact: bool = False,
) -> bool:
    base_set = extract_formulas(base_file, ignore_sheets=ignore_sheets, only_sheets=only_sheets)
    target_set = extract_formulas(target_file, ignore_sheets=ignore_sheets, only_sheets=only_sheets)

    diff_lines = list(
        difflib.unified_diff(
            base_set,
            target_set,
            fromfile=base_file,
            tofile=target_file,
            lineterm="",
            n=0,  # no context lines — ever
        )
    )

    if not diff_lines:
        msg = "FormDiff complete: zero formula deviations detected."
        if output_path:
            Path(output_path).write_text(msg + "\n", encoding="utf-8")
            html_path = _html_output_path(output_path)
            Path(html_path).write_text(
                f"<html><body><p style='font-family:monospace;color:green'>{msg}</p></body></html>",
                encoding="utf-8",
            )
        else:
            print(Fore.GREEN + msg)
        return False

    if output_path:
        Path(output_path).write_text("\n".join(diff_lines) + "\n", encoding="utf-8")
        print(f"Diff written to {output_path}")

        if compact:
            compact_lines = build_compact_diff(diff_lines)
            compact_path = _compact_output_path(output_path)
            Path(compact_path).write_text("\n".join(compact_lines) + "\n", encoding="utf-8")
            print(f"Compact diff written to {compact_path}")
            compact_html = _build_html_report(compact_lines, base_file, target_file)
            compact_html_path = _html_output_path(compact_path)
            Path(compact_html_path).write_text(compact_html, encoding="utf-8")
            print(f"Compact HTML report written to {compact_html_path}")

        html_report = _build_html_report(diff_lines, base_file, target_file)
        html_path = _html_output_path(output_path)
        Path(html_path).write_text(html_report, encoding="utf-8")
        print(f"HTML report written to {html_path}")
    else:
        display_lines = build_compact_diff(diff_lines) if compact else diff_lines
        for line in display_lines:
            if line.startswith("+++") or line.startswith("---"):
                print(Style.BRIGHT + line)
            elif line.startswith("+"):
                print(Fore.GREEN + line)
            elif line.startswith("-"):
                print(Fore.RED + line)
            elif line.startswith("@@"):
                print(Fore.CYAN + line)
            else:
                print(line)

    if summary:
        print_summary(diff_lines)

    return True


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="formdiff",
        description="Git-like formula diff for Excel workbooks.",
    )
    parser.add_argument("base", help="Path to the base (old) .xlsx workbook")
    parser.add_argument("target", help="Path to the target (new) .xlsx workbook")
    parser.add_argument(
        "-o",
        "--output",
        metavar="FILE",
        default=None,
        help="Write diff to FILE (.diff) and auto-generate FILE.html report alongside it",
    )
    parser.add_argument(
        "-x",
        "--exclude-sheet",
        action="append",
        metavar="SHEET",
        dest="exclude_sheets",
        default=[],
        help="Sheet name to ignore (can be repeated: -x LTOP -x Summary)",
    )
    parser.add_argument(
        "-i",
        "--include-sheet",
        action="append",
        metavar="SHEET",
        dest="include_sheets",
        default=[],
        help="Only diff this sheet — all others are ignored (can be repeated: -i BOM -i F2)",
    )
    parser.add_argument(
        "-s",
        "--summary",
        action="store_true",
        default=False,
        help="Print a grouped change summary after the diff",
    )
    parser.add_argument(
        "--compact",
        action="store_true",
        default=False,
        help="Collapse runs of identical consecutive-row changes. With -o, writes a separate FILE-compact.diff",
    )
    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    for path in (args.base, args.target):
        if not Path(path).is_file():
            print(f"File not found: {path}", file=sys.stderr)
            sys.exit(1)

    ignore_sheets = set(args.exclude_sheets) if args.exclude_sheets else None
    only_sheets = set(args.include_sheets) if args.include_sheets else None

    if only_sheets:
        print(f"Only diffing sheets: {', '.join(sorted(only_sheets))}")
    if ignore_sheets:
        print(f"Ignoring sheets: {', '.join(sorted(ignore_sheets))}")

    print(f"FormDiff: {args.base} → {args.target}\n")
    changed = run_diff(
        args.base,
        args.target,
        output_path=args.output,
        ignore_sheets=ignore_sheets,
        only_sheets=only_sheets,
        summary=args.summary,
        compact=args.compact,
    )
    sys.exit(1 if changed else 0)


if __name__ == "__main__":
    main()
